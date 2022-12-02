import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import os

# filename = 'data/Приложение_№_1_Чеченская_Республика_01_12.xlsx'
# filename = 'data/Ингушетия Приложение_№_1.xlsx'
dir_name = 'data'

# Создаем итоговый файл excel
wb = openpyxl.Workbook()
# создаем листы
ren_sheet = wb['Sheet']
ren_sheet.title = 'Свод-студенты'
wb.create_sheet(title='Свод-кадры', index=1)
wb.create_sheet(title='Свод-финансы и МТБ', index=2)

# Создаем итоговые датафреймы
students_df = pd.DataFrame()
kadr_df = pd.DataFrame()
fin_df = pd.DataFrame()
for dirpath, dirnames, filenames in os.walk(dir_name):
    for filename in filenames:
        if filename.endswith('.xlsx'):
            # Получаем название файла без расширения
            name_file = filename.split('.xlsx')[0]
            print(name_file)
            temb_wb = load_workbook(filename=f'{dirpath}/{filename}', read_only=True)  # загружаем файл
            students_name_sheet = 'Не найден лист с таким названием'
            kadr_name_sheet = 'Не найден лист с таким названием'
            fin_name_sheet = 'Не найден лист с таким названием'

            for name_sheet in temb_wb.sheetnames:  # получаем названия листов, поскольку они могут быть с большой или маленькой буквы
                if 'денты' in name_sheet:
                    students_name_sheet = name_sheet
                if 'адры' in name_sheet:
                    kadr_name_sheet = name_sheet
                if 'ансы' in name_sheet:
                    fin_name_sheet = name_sheet
                else:
                    continue
            # Создаем датафреймы
            temp_stud_df = pd.read_excel(filename, sheet_name=students_name_sheet)


            temp_kadr_df = pd.read_excel(filename, sheet_name=kadr_name_sheet)


            temp_fin_df = pd.read_excel(filename, sheet_name=fin_name_sheet)
            # Добавляем колонку с названием файла откуда взяты данные
            temp_stud_df.insert(0, 'Откуда взяты данные', filename)
            temp_kadr_df.insert(0, 'Откуда взяты данные', filename)
            temp_fin_df.insert(0, 'Откуда взяты данные', filename)
            students_df = pd.concat([students_df,temp_stud_df],ignore_index=True)
            kadr_df = pd.concat([kadr_df,temp_kadr_df],ignore_index=True)
            fin_df = pd.concat([fin_df,temp_fin_df],ignore_index=True)


# записываем в соответствующий лист
for r in dataframe_to_rows(students_df, index=False, header=False):
    wb['Свод-студенты'].append(r)
for r in dataframe_to_rows(kadr_df, index=False, header=False):
    wb['Свод-кадры'].append(r)
for r in dataframe_to_rows(fin_df, index=False, header=False):
    wb['Свод-финансы и МТБ'].append(r)

t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл
wb.save(f'Свод от {current_time}.xlsx')