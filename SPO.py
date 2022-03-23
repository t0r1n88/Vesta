
import pandas as pd
import os
# from docxtpl import DocxTemplate
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time
import datetime
from datetime import date
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart, Reference, PieChart, PieChart3D, Series
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
import sys
import locale


def set_rus_locale():
    """
    Функция чтобы можно было извлечь русские названия месяцев
    """
    locale.setlocale(
        locale.LC_ALL,
        'rus_rus' if sys.platform == 'win32' else 'ru_RU.UTF-8')


def calculate_age(born):
    """
    Функция для расчета текущего возраста взято с https://stackoverflow.com/questions/2217488/age-from-birthdate-in-python/9754466#9754466
    :param born: дата рождения
    :return: возраст
    """

    try:
        today = date.today()
        return today.year - born.year - ((today.month, today.day) < (born.month, born.day))
    except:
        print(born)
        messagebox.showerror('ЦОПП Бурятия', 'Отсутствует или некорректная дата рождения слушателя\nПроверьте файл!')
        quit()


def convert_date(cell):
    """
    Функция для конвертации даты в формате 1957-05-10 в формат 10.05.1957(строковый)
    """

    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except TypeError:
        print(cell)
        messagebox.showerror('ЦОПП Бурятия', 'Проверьте правильность заполнения ячеек с датой!!!')
        quit()

def extract_number_month(cell):
    """
    Функция для извлечения номера месяца
    """
    return cell.month


def extract_name_month(cell):
    """
    Функция для извлечения названия месяца
    Взято отсюда https://ru.stackoverflow.com/questions/1045154/Вывод-русских-символов-из-pd-timestamp-month-name
    """
    return cell.month_name(locale='Russian')


def extract_year(cell):
    """
    Функция для извлечения года рождения
    """
    return cell.year


# Устанавливаем русскую локаль
set_rus_locale()

name_file_data_date = 'Тест.xlsx'

path_to_end_folder_date = 'data/'

name_column = 'Дата_рождения_получателя'

# Считываем файл
df = pd.read_excel(name_file_data_date)
# Конвертируем его в формат даты
df[name_column] = pd.to_datetime(df[name_column],dayfirst=True)
# Создаем шрифт которым будем выделять названия таблиц
font_name_table = Font(name='Arial Black', size=15, italic=True)

# Создаем файл excel
wb = openpyxl.Workbook()
# Создаем листы
wb.create_sheet(title='Итоговая таблица', index=0)
wb.create_sheet(title='Свод по возрастам', index=1)
wb.create_sheet(title='Свод по месяцам', index=2)
wb.create_sheet(title='Свод по годам', index=3)
wb.create_sheet(title='Свод по 1-ПК', index=4)
wb.create_sheet(title='Свод по 1-ПО', index=5)
wb.create_sheet(title='Свод по СПО-1', index=6)
wb.create_sheet(title='Свод по категориям Росстата', index=7)



# Подсчитываем текущий возраст
df['Текущий возраст'] = df[name_column].apply(calculate_age)

# Получаем номер месяца
df['Порядковый номер месяца рождения'] = df[name_column].apply(extract_number_month)

# Получаем название месяца
df['Название месяца рождения'] = df[name_column].apply(extract_name_month)

# Получаем год рождения
df['Год рождения'] = df[name_column].apply(extract_year)

# Присваиваем категорию по 1-ПК
df['1-ПК Категория'] = pd.cut(df['Текущий возраст'], [0, 24, 29, 34, 39, 44, 49, 54, 59, 64, 101, 10000],
                              labels=['моложе 25 лет', '25-29', '30-34', '35-39',
                                      '40-44', '45-49', '50-54', '55-59', '60-64',
                                      '65 и более',
                                      'Возраст  больше 101'])
# Приводим к строковому виду, иначе не запишется на лист
df['1-ПК Категория'] = df['1-ПК Категория'].astype(str)

# Присваиваем категорию по 1-ПО
df['1-ПО Категория'] = pd.cut(df['Текущий возраст'],
                              [0, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25,
                               26, 27, 28,
                               29, 34, 39, 44, 49, 54, 59, 64, 101],
                              labels=['моложе 14 лет', '14 лет', '15 лет',
                                      '16 лет',
                                      '17 лет', '18 лет', '19 лет', '20 лет',
                                      '21 год', '22 года',
                                      '23 года', '24 года', '25 лет',
                                      '26 лет', '27 лет', '28 лет', '29 лет',
                                      '30-34 лет',
                                      '35-39 лет', '40-44 лет', '45-49 лет',
                                      '50-54 лет',
                                      '55-59 лет',
                                      '60-64 лет',
                                      '65 лет и старше'])
# Приводим к строковому виду, иначе не запишется на лист
df['1-ПО Категория'] = df['1-ПО Категория'].astype(str)

# Присваиваем категорию по 1-СПО
df['СПО-1 Категория'] = pd.cut(df['Текущий возраст'],
                               [0, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 34, 39, 101],
                               labels=['моложе 13 лет', '13 лет', '14 лет', '15 лет', '16 лет', '17 лет', '18 лет',
                                       '19 лет', '20 лет'
                                   , '21 год', '22 года', '23 года', '24 года', '25 лет', '26 лет', '27 лет', '28 лет',
                                       '29 лет',
                                       '30-34 лет', '35-39 лет', '40 лет и старше'])
## Приводим к строковому виду, иначе не запишется на лист
df['СПО-1 Категория'] = df['СПО-1 Категория'].astype(str)

# Присваиваем категорию по Росстату
df['Росстат Категория'] = pd.cut(df['Текущий возраст'], [0, 4, 9, 14, 19, 24, 29, 34, 39, 44, 49, 54, 59, 64, 69, 200],
                                 labels=['0-4', '5-9', '10-14', '15-19', '20-24', '25-29', '30-34',
                                         '35-39', '40-44', '45-49', '50-54', '55-59', '60-64', '65-69',
                                         '70 лет и старше'])
## Приводим к строковому виду, иначе не запишется на лист
df['Росстат Категория'] = df['Росстат Категория'].astype(str)

# заполняем сводные таблицы
# Сводная по возрастам

df_svod_by_age = df.groupby(['Текущий возраст']).agg({name_column: 'count'})
df_svod_by_age.columns = ['Количество']

for r in dataframe_to_rows(df_svod_by_age, index=True, header=True):
    wb['Свод по возрастам'].append(r)

# Сводная по месяцам
df_svod_by_month = df.groupby(['Название месяца рождения']).agg({name_column: 'count'})
df_svod_by_month.columns = ['Количество']

# Сортируем индекс чтобы месяцы шли в хоронологическом порядке
# Взял отсюда https://stackoverflow.com/questions/40816144/pandas-series-sort-by-month-index
df_svod_by_month.index = pd.CategoricalIndex(df_svod_by_month.index,
                                             categories=['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль',
                                                         'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'],
                                             ordered=True)
df_svod_by_month.sort_index(inplace=True)

for r in dataframe_to_rows(df_svod_by_month, index=True, header=True):
    wb['Свод по месяцам'].append(r)

# Сводная по годам
df_svod_by_year = df.groupby(['Год рождения']).agg({name_column: 'count'})
df_svod_by_year.columns = ['Количество']

for r in dataframe_to_rows(df_svod_by_year, index=True, header=True):
    wb['Свод по годам'].append(r)

# Сводная по 1-ПК
df_svod_by_1PK = df.groupby(['1-ПК Категория']).agg({name_column: 'count'})
df_svod_by_1PK.columns = ['Количество']

for r in dataframe_to_rows(df_svod_by_1PK, index=True, header=True):
    wb['Свод по 1-ПК'].append(r)

# Сводная по 1-ПО
df_svod_by_1PO = df.groupby(['1-ПО Категория']).agg({name_column: 'count'})
df_svod_by_1PO.columns = ['Количество']

for r in dataframe_to_rows(df_svod_by_1PO, index=True, header=True):
    wb['Свод по 1-ПО'].append(r)

# Сводная по СПО-1
df_svod_by_SPO1 = df.groupby(['СПО-1 Категория']).agg({name_column: 'count'})
df_svod_by_SPO1.columns = ['Количество']

for r in dataframe_to_rows(df_svod_by_SPO1, index=True, header=True):
    wb['Свод по СПО-1'].append(r)

# Сводная по Росстату
df_svod_by_Ros = df.groupby(['Росстат Категория']).agg({name_column: 'count'})
df_svod_by_Ros.columns = ['Количество']

# Сортируем индекс
df_svod_by_Ros.index = pd.CategoricalIndex(df_svod_by_Ros.index,
                                           categories=['0-4', '5-9', '10-14', '15-19', '20-24', '25-29', '30-34',
                                                       '35-39', '40-44', '45-49', '50-54', '55-59', '60-64', '65-69',
                                                       '70 лет и старше', 'nan'],
                                           ordered=True)
df_svod_by_Ros.sort_index(inplace=True)

for r in dataframe_to_rows(df_svod_by_Ros, index=True, header=True):
    wb['Свод по категориям Росстата'].append(r)

for r in dataframe_to_rows(df, index=False, header=True):
    wb['Итоговая таблица'].append(r)

t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)
# Сохраняем итоговый файл
wb.save(f'{path_to_end_folder_date}/Таблица по датам рождения от {current_time}.xlsx')

