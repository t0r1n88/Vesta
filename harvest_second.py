"""
Скрипт для обработки данных  с многостраничных листов Excel
"""
import pandas as pd
import os
from dateutil.parser import ParserError
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import time
import datetime
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
pd.options.mode.chained_assignment = None

skip_rows = 1
# file_standard_merger = 'data/harvest/Приложение_№_1_Чеченская_Республика_01_12 (2).xlsx'
# file_standard_merger = 'data/harvest/Ингушетия Приложение_№_1 (2).xlsx'
file_standard_merger = 'data/union/Ингушетия Приложение_№_1 (2).xlsx'
# file_standard_merger = 'data/temp2/Список 24.05.01 Проектирование, производство и эксплуатация ракет и ракетно-космических комплексов.xlsx'
# dir_name = 'data/harvest'
dir_name = 'data/union'
# dir_name = 'data/temp2'
path_to_end_folder_merger = 'data/temp'
params_harvest = 'data/params.xlsx'  # файл с параметрами
checkbox_harvest = 2

# Создаем датафрейм куда будем сохранять ошибочные файлы
err_df = pd.DataFrame(columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'])

name_file_standard_merger = file_standard_merger.split('/')[-1]  # получаем имя файла

standard_wb = load_workbook(filename=file_standard_merger)  # Загружаем эталонный файл

standard_sheets = sorted(standard_wb.sheetnames)  # отсортрованный список листов по которому будет вестись сравнение
set_standard_sheets = set(standard_sheets)  # создаем множество из листов эталонного файла
standard_size_sheets = len(standard_sheets)

dct_df = dict()  # создаем словарь в котором будем хранить да

for sheet in standard_wb.sheetnames:  # Добавляем в словарь датафреймы
    temp_df = pd.read_excel(file_standard_merger, sheet_name=sheet, dtype=str)
    dct_df[sheet] = temp_df

if checkbox_harvest == 0:  # Вариант объединения по названию листов
    for dirpath, dirnames, filenames in os.walk(dir_name):
        for filename in filenames:
            if (filename.endswith('.xlsx') and not filename.startswith(
                    '~$')) and filename != name_file_standard_merger:  # не обрабатываем эталонный файл
                # Получаем название файла без расширения
                name_file = filename.split('.xlsx')[0]
                print(name_file)
                temb_wb = load_workbook(filename=f'{dirpath}/{filename}')  # загружаем файл, для проверки листов
                """
                Проверяем наличие листов из эталонного файла в проверяемом файле, если они есть то начинаем 
                дальнейшую проверку
                """
                if set_standard_sheets.issubset(set(temb_wb.sheetnames)):
                    count_errors = 0
                    # проверяем наличие листов указанных в файле параметров
                    for name_sheet, df in dct_df.items():  # Проводим проверку на совпадение
                        lst_df = pd.read_excel(f'{dirpath}/{filename}',sheet_name=name_sheet)
                        if lst_df.shape[1] != df.shape[1]:
                            # если количество колонок не совпадает то записываем как ошибку
                            temp_error_df = pd.DataFrame(
                                columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                                data=[[name_file, name_sheet, 'Количество колонок отличается от эталонного',
                                       f'Ожидалось {df.shape[1]} колонок, а в листе {lst_df.shape[1]}']])  # создаем временный датафрейм. потом надо подумать над словарем

                            err_df = pd.concat([err_df, temp_error_df],
                                               ignore_index=True)  # добавляем в датафрейм ошибок
                            count_errors += 1

                    # если хоть одна ошибка то проверяем следующий файл
                    if count_errors != 0:
                        continue
                    # если нет то начинаем обрабатывать листы
                    for name_sheet, df in dct_df.items():
                        temp_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=name_sheet,
                                                dtype=str)  # загружаем датафрейм
                        temp_df['Откуда взяты данные'] = name_file
                        for row in dataframe_to_rows(temp_df, index=False, header=False):
                            standard_wb[name_sheet].append(row)  # добавляем данные

                elif len(temb_wb.sheetnames) == standard_size_sheets:  # сравниваем количество листов в файле
                    diff_name_sheets = set(temb_wb.sheetnames).difference(
                        set(standard_sheets))  # проверяем разницу в названиях листов
                    print(diff_name_sheets)
                    if len(diff_name_sheets) != 0:  # если разница в названиях есть то записываем в ошибки и обрабатываем следующий файл
                        temp_error_df = pd.DataFrame(
                            columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'], data=[
                                [name_file, '', 'Названия листов отличаются от эталонных',
                                 f'Отличаются следующие названия листов {diff_name_sheets}']])  # создаем временный датафрейм. потом надо подумать над словарем

                        err_df = pd.concat([err_df, temp_error_df], ignore_index=True)  # добавляем в датафрейм ошибок

                        continue

                else:
                    temp_error_df = pd.DataFrame(
                        columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                        data=[[name_file,'', 'Не совпадает количество или название листов в файле',
                               f'Листы, которые есть в файле: {",".join(temb_wb.sheetnames)}']])  # создаем временный датафрейм. потом надо подумать над словарем

                    err_df = pd.concat([err_df, temp_error_df],
                                       ignore_index=True)  # добавляем в датафрейм ошибок

    # Получаем текущую дату
    current_time = time.strftime('%H_%M_%S %d.%m.%Y')
    standard_wb.save(f'{path_to_end_folder_merger}/Слияние по варианту А Общая таблица от {current_time}.xlsx')  # сохраняем
    err_df.to_excel(f'{path_to_end_folder_merger}/Слияние по варианту А Ошибки от {current_time}.xlsx',
                    index=False)  # сохраняем ошибки
elif checkbox_harvest == 1:  # Вариант объединения по порядку
    for dirpath, dirnames, filenames in os.walk(dir_name):
        for filename in filenames:
            if (filename.endswith('.xlsx') and not filename.startswith(
                    '~$')) and filename != name_file_standard_merger:  # не обрабатываем эталонный файл
                # Получаем название файла без расширения
                name_file = filename.split('.xlsx')[0]
                print(name_file)
                temb_wb = load_workbook(filename=f'{dirpath}/{filename}')  # загружаем файл

                if standard_size_sheets == len(temb_wb.sheetnames):  # если количество листов одинаково то обрабатываем
                    count_errors = 0  # счетчик ошибок
                    dct_name_sheet = {}  # создаем словарь где ключ это название листа в эталонном файле а значение это название листа в обрабатываемом файле
                    for idx, data in enumerate(dct_df.items()):  # Проводим проверку на совпадение
                        name_sheet = data[0]  # получаем название листа
                        df = data[1]  # получаем датафрейм
                        temp_name_sheet = temb_wb.sheetnames[idx]  #
                        lst_df = pd.read_excel(f'{dirpath}/{filename}',sheet_name=temp_name_sheet)
                        if lst_df.shape[1] != df.shape[1]:
                            # если количество колонок не совпадает то записываем как ошибку
                            temp_error_df = pd.DataFrame(
                                columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                                data=[
                                    [name_file, name_sheet, 'Количество колонок отличается от эталонного',
                                     f'Ожидалось {df.shape[1]} колонок, а в листе {lst_df.shape[1]}']])  # создаем временный датафрейм. потом надо подумать над словарем

                            err_df = pd.concat([err_df, temp_error_df],
                                               ignore_index=True)  # добавляем в датафрейм ошибок
                            count_errors += 1

                        else:
                            dct_name_sheet[name_sheet] = temp_name_sheet
                    # если хоть одна ошибка то проверяем следующий файл
                    if count_errors != 0:
                        continue
                        # если нет то начинаем обрабатывать листы
                    for name_sheet, df in dct_df.items():
                        temp_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=dct_name_sheet[name_sheet],
                                                dtype=str)  # загружаем датафрейм
                        temp_df['Откуда взяты данные'] = name_file
                        for row in dataframe_to_rows(temp_df, index=False, header=False):
                            standard_wb[name_sheet].append(row)  # добавляем данные
                else:
                    temp_error_df = pd.DataFrame(
                        columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                        data=[[name_file, '', 'Не совпадает количество или название листов в файле',
                               f'Листы, которые есть в файле: {",".join(temb_wb.sheetnames)}']])  # создаем временный датафрейм. потом надо подумать над словарем

                    err_df = pd.concat([err_df, temp_error_df],
                                       ignore_index=True)  # добавляем в датафрейм ошибок

    # Получаем текущую дату
    current_time = time.strftime('%H_%M_%S %d.%m.%Y')
    standard_wb.save(f'{path_to_end_folder_merger}/Слияние по варианту Б Общая таблица от {current_time}.xlsx')  # сохраняем

    err_df.to_excel(f'{path_to_end_folder_merger}/Слияние по варианту Б Ошибки о {current_time}.xlsx', index=False)

# Если выбран управляемый сбор данных
elif checkbox_harvest == 2:
    df_params = pd.read_excel(params_harvest, header=None)  # загружаем параметры
    print(df_params)
    tmp_name_sheets = df_params[0].tolist()  # создаем списки чтобы потом из них сделать словарь
    tmp_skip_rows = df_params[1].tolist()
    dct_manage_harvest = dict(zip(tmp_name_sheets,
                                  tmp_skip_rows))  # создаем словарь где ключ это название листа а значение это сколько строк нужно пропустить
    print(dct_manage_harvest)
    set_params_sheets = set(dct_manage_harvest.keys())  # создаем множество из ключей(листов) которые нужно обработать
    if not set_params_sheets.issubset(
            set_standard_sheets):  # проверяем совпадение названий в эталонном файле и в файле параметров
        diff_value = set(dct_manage_harvest.keys()).difference(set(standard_sheets))  # получаем разницу

        messagebox.showerror('', f'Не совпадают следующие названия листов в файле параметров и в эталонном файле\n'
                                 f'{diff_value}!')
    # начинаем обработку
    for dirpath, dirnames, filenames in os.walk(dir_name):
        for filename in filenames:
            if (filename.endswith('.xlsx') and not filename.startswith(
                    '~$')) and filename != name_file_standard_merger:  # не обрабатываем эталонный файл
                # Получаем название файла без расширения
                name_file = filename.split('.xlsx')[0]
                # print(name_file)
                temb_wb = load_workbook(filename=f'{dirpath}/{filename}')  # загружаем файл
                if set_params_sheets.issubset(set(temb_wb.sheetnames)):
                    count_errors = 0
                    # проверяем наличие листов указанных в файле параметров
                    for name_sheet, skip_r in dct_manage_harvest.items():  # Проводим проверку на совпадение количества колонок
                        lst_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=name_sheet)
                        if lst_df.shape[1] != dct_df[name_sheet].shape[1]:
                            # если количество колонок не совпадает то записываем как ошибку
                            temp_error_df = pd.DataFrame(
                                columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                                data=[[name_file, name_sheet, 'Количество колонок отличается от эталонного',
                                       f'Ожидалось {dct_df[name_sheet].shape[1]} колонок, а в листе {lst_df.shape[1]}']])  # создаем временный датафрейм. потом надо подумать над словарем
                            err_df = pd.concat([err_df, temp_error_df],
                                               ignore_index=True)  # добавляем в датафрейм ошибок
                            count_errors += 1
                    #
                    # если хоть одна ошибка то проверяем следующий файл
                    if count_errors != 0:
                        continue
                    # если нет то начинаем обрабатывать листы
                    for name_sheet, skip_r in dct_manage_harvest.items():
                        temp_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=name_sheet,skiprows=skip_r,
                                                dtype=str,header=None)  # загружаем датафрейм
                        temp_df['Откуда взяты данные'] = name_file
                        for row in dataframe_to_rows(temp_df, index=False, header=False):
                            standard_wb[name_sheet].append(row)  # добавляем данные
                else:
                    temp_error_df = pd.DataFrame(
                        columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                        data=[[name_file, '', 'Не совпадает количество или название листов в файле',
                               f'Листы, которые есть в файле: {",".join(temb_wb.sheetnames)}']])  # создаем временный датафрейм. потом надо подумать над словарем

                    err_df = pd.concat([err_df, temp_error_df],
                                       ignore_index=True)  # добавляем в датафрейм ошибок

    # # Получаем текущую дату
    current_time = time.strftime('%H_%M_%S %d.%m.%Y')
    standard_wb.save(f'{path_to_end_folder_merger}/Слияние по варианту В Общая таблица от {current_time}.xlsx')  # сохраняем
    err_df.to_excel(f'{path_to_end_folder_merger}/Слияние по варианту В Ошибки от {current_time}.xlsx',
                    index=False)  # сохраняем ошибки
    #
