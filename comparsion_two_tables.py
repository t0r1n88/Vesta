"""
Функции для соединения, сравнения, слияния 2 таблиц
"""
import pandas as pd
import numpy as np

from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import datetime
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging

logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)


def convert_columns_to_str(df, number_columns):
    """
    Функция для конвертации указанных столбцов в строковый тип и очистки от пробельных символов в начале и конце
    """

    for column in number_columns:  # Перебираем список нужных колонок
        try:
            df.iloc[:, column] = df.iloc[:, column].astype(str)
            # Очищаем колонку от пробельных символов с начала и конца
            df.iloc[:, column] = df.iloc[:, column].apply(lambda x: x.strip())
        except IndexError:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 'Проверьте порядковые номера колонок которые вы хотите обработать.')


def convert_params_columns_to_int(lst):
    """
    Функция для конвератации значений колонок которые нужно обработать.
    Очищает от пустых строк, чтобы в итоге остался список из чисел в формате int
    """
    out_lst = []  # Создаем список в который будем добавлять только числа
    for value in lst:  # Перебираем список
        try:
            # Обрабатываем случай с нулем, для того чтобы после приведения к питоновскому отсчету от нуля не получилась колонка с номером -1
            number = int(value)
            if number != 0:
                out_lst.append(value)  # Если конвертирования прошло без ошибок то добавляем
            else:
                continue
        except:  # Иначе пропускаем
            continue
    return out_lst


def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'
    except TypeError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'


def clean_ending_columns(lst_columns: list, name_first_df, name_second_df):
    """
    Функция для очистки колонок таблицы с совпадающими данными от окончаний _x _y

    :param lst_columns:
    :param name_first_df
    :param name_second_df
    :return:
    """
    out_columns = []  # список для очищенных названий
    for name_column in lst_columns:
        if '_x' in name_column:
            # если они есть то проводим очистку и добавление времени
            cut_name_column = name_column[:-2]  # обрезаем
            temp_name = f'{cut_name_column}_{name_first_df}'  # соединяем
            out_columns.append(temp_name)  # добавляем
        elif '_y' in name_column:
            cut_name_column = name_column[:-2]  # обрезаем
            temp_name = f'{cut_name_column}_{name_second_df}'  # соединяем
            out_columns.append(temp_name)  # добавляем
        else:
            out_columns.append(name_column)
    return out_columns


def merging_two_tables(file_params, first_sheet_name, second_sheet_name, first_file, second_file,
                       path_to_end_folder_comparison):
    """
    Функция для сравнения и слияния 2 таблиц
    :param file_params: файл с порядковыми номерами колонок по которым нужно произвести слияние
    :param first_sheet_name: имя листа в первом файле
    :param second_sheet_name: имя листа во втором файле
    :param first_file: путь к первому файлу
    :param second_file: путь ко второму файлу
    :param path_to_end_folder_comparison: путь к итоговой папке
    :return:
    """
    try:
        # загружаем файлы
        first_df = pd.read_excel(first_file, sheet_name=first_sheet_name, dtype=str,
                                 keep_default_na=False)
        # получаем имя файла
        name_first_df = first_file.split('/')[-1]
        name_first_df = name_first_df.split('.xlsx')[0]

        second_df = pd.read_excel(second_file, sheet_name=second_sheet_name, dtype=str,
                                  keep_default_na=False)
        # получаем имя файла
        name_second_df = second_file.split('/')[-1]
        name_second_df = name_second_df.split('.xlsx')[0]

        params = pd.read_excel(file_params, header=None, keep_default_na=False)

        # Преврашаем каждую колонку в список
        params_first_columns = params[0].tolist()
        params_second_columns = params[1].tolist()

        # Конвертируем в инт заодно проверяя корректность введенных данных
        int_params_first_columns = convert_params_columns_to_int(params_first_columns)
        int_params_second_columns = convert_params_columns_to_int(params_second_columns)

        # Отнимаем 1 от каждого значения чтобы привести к питоновским индексам
        int_params_first_columns = list(map(lambda x: x - 1, int_params_first_columns))
        int_params_second_columns = list(map(lambda x: x - 1, int_params_second_columns))

        # Конвертируем нужные нам колонки в str
        convert_columns_to_str(first_df, int_params_first_columns)
        convert_columns_to_str(second_df, int_params_second_columns)

        # Проверяем наличие колонок с датами в списке колонок для объединения чтобы привести их в нормальный вид
        for number_column_params in int_params_first_columns:
            if 'дата' in first_df.columns[number_column_params].lower():
                first_df.iloc[:, number_column_params] = pd.to_datetime(first_df.iloc[:, number_column_params],
                                                                        errors='coerce', dayfirst=True)
                first_df.iloc[:, number_column_params] = first_df.iloc[:, number_column_params].apply(
                    create_doc_convert_date)

        for number_column_params in int_params_second_columns:
            if 'дата' in second_df.columns[number_column_params].lower():
                second_df.iloc[:, number_column_params] = pd.to_datetime(second_df.iloc[:, number_column_params],
                                                                         errors='coerce', dayfirst=True)
                second_df.iloc[:, number_column_params] = second_df.iloc[:, number_column_params].apply(
                    create_doc_convert_date)

        # в этом месте конвертируем даты в формат ДД.ММ.ГГГГ
        # processing_date_column(first_df, int_params_first_columns)
        # processing_date_column(second_df, int_params_second_columns)

        # Проверяем наличие колонки _merge
        if '_merge' in first_df.columns:
            first_df.drop(columns=['_merge'], inplace=True)
        if '_merge' in second_df.columns:
            second_df.drop(columns=['_merge'], inplace=True)
        # Проверяем наличие колонки ID
        if 'ID_объединения' in first_df.columns:
            first_df.drop(columns=['ID_объединения'], inplace=True)
        if 'ID_объединения' in second_df.columns:
            second_df.drop(columns=['ID_объединения'], inplace=True)

        # создаем датафреймы из колонок выбранных для объединения, такой способо связан с тем, что
        # при использовании sum числа в строковом виде превращаются в числа
        key_first_df = first_df.iloc[:, int_params_first_columns]
        key_second_df = second_df.iloc[:, int_params_second_columns]
        # Создаем в каждом датафрейме колонку с айди путем склеивания всех нужных колонок в одну строку
        first_df['ID_объединения'] = key_first_df.apply(lambda x: ''.join(x), axis=1)
        second_df['ID_объединения'] = key_second_df.apply(lambda x: ''.join(x), axis=1)

        first_df['ID_объединения'] = first_df['ID_объединения'].apply(lambda x: x.replace(' ', ''))
        second_df['ID_объединения'] = second_df['ID_объединения'].apply(lambda x: x.replace(' ', ''))

        # делаем прописными айди значения по которым будет вестись объединение
        first_df['ID_объединения'] = first_df['ID_объединения'].apply(lambda x: x.upper())
        second_df['ID_объединения'] = second_df['ID_объединения'].apply(lambda x: x.upper())

        # В результат объединения попадают совпадающие по ключу записи обеих таблиц и все строки из этих двух таблиц, для которых пар не нашлось. Порядок таблиц в запросе не

        # Создаем документ
        wb = openpyxl.Workbook()
        # создаем листы
        ren_sheet = wb['Sheet']
        ren_sheet.title = 'Таблица 1'
        wb.create_sheet(title='Таблица 2', index=1)
        wb.create_sheet(title='Совпадающие данные', index=2)
        wb.create_sheet(title='Обновленная таблица', index=3)
        wb.create_sheet(title='Объединённая таблица', index=4)

        # Создаем переменные содержащие в себе количество колонок в базовых датареймах
        first_df_quantity_cols = len(first_df.columns)  # не забываем что там добавилась колонка ID

        # Проводим слияние
        itog_df = pd.merge(first_df, second_df, how='outer', left_on=['ID_объединения'], right_on=['ID_объединения'],
                           indicator=True)

        # копируем в отдельный датафрейм для создания таблицы с обновлениями
        update_df = itog_df.copy()

        # Записываем каждый датафрейм в соответсвующий лист
        # Левая таблица
        left_df = itog_df[itog_df['_merge'] == 'left_only']
        left_df.drop(['_merge'], axis=1, inplace=True)

        # Удаляем колонки второй таблицы чтобы не мешались
        left_df.drop(left_df.iloc[:, first_df_quantity_cols:], axis=1, inplace=True)

        # Переименовываем колонки у которых были совпадение во второй таблице, в таких колонках есть добавление _x
        clean_left_columns = list(map(lambda x: x[:-2] if '_x' in x else x, list(left_df.columns)))
        left_df.columns = clean_left_columns
        for r in dataframe_to_rows(left_df, index=False, header=True):
            wb['Таблица 1'].append(r)

        right_df = itog_df[itog_df['_merge'] == 'right_only']
        right_df.drop(['_merge'], axis=1, inplace=True)

        # Удаляем колонки первой таблицы таблицы чтобы не мешались
        right_df.drop(right_df.iloc[:, :first_df_quantity_cols - 1], axis=1, inplace=True)

        # Переименовываем колонки у которых были совпадение во второй таблице, в таких колонках есть добавление _x
        clean_right_columns = list(map(lambda x: x[:-2] if '_y' in x else x, list(right_df.columns)))
        right_df.columns = clean_right_columns

        for r in dataframe_to_rows(right_df, index=False, header=True):
            wb['Таблица 2'].append(r)

        both_df = itog_df[itog_df['_merge'] == 'both']
        both_df.drop(['_merge'], axis=1, inplace=True)
        # Очищаем от _x  и _y
        clean_both_columns = clean_ending_columns(list(both_df.columns), name_first_df, name_second_df)
        both_df.columns = clean_both_columns

        for r in dataframe_to_rows(both_df, index=False, header=True):
            wb['Совпадающие данные'].append(r)

        # Сохраняем общую таблицу
        # Заменяем названия индикаторов на более понятные
        itog_df['_merge'] = itog_df['_merge'].apply(lambda x: 'Данные из первой таблицы' if x == 'left_only' else
        ('Данные из второй таблицы' if x == 'right_only' else 'Совпадающие данные'))
        itog_df['_merge'] = itog_df['_merge'].astype(str)

        clean_itog_df = clean_ending_columns(list(itog_df.columns), name_first_df, name_second_df)
        itog_df.columns = clean_itog_df
        for r in dataframe_to_rows(itog_df, index=False, header=True):
            wb['Объединённая таблица'].append(r)

        # получаем список с совпадающими колонками первой таблицы
        first_df_columns = [column for column in list(update_df.columns) if str(column).endswith('_x')]
        # получаем список с совпадающими колонками второй таблицы
        second_df_columns = [column for column in list(update_df.columns) if str(column).endswith('_y')]
        # Создаем из списка совпадающих колонок второй таблицы словарь, чтобы было легче обрабатывать
        # да конечно можно было сделать в одном выражении но как я буду читать это через 2 недели?
        dct_second_columns = {column.split('_y')[0]: column for column in second_df_columns}

        for column in first_df_columns:
            # очищаем от _x
            name_column = column.split('_x')[0]
            # Обновляем значение в случае если в колонке _merge стоит both, иначе оставляем старое значение,
            # Чтобы обновить значение в ячейке, во второй таблице не должно быть пустого значения или пробела в аналогичной колонке

            update_df[column] = np.where(
                (update_df['_merge'] == 'both') & (update_df[dct_second_columns[name_column]]) & (
                        update_df[dct_second_columns[name_column]] != ' '),
                update_df[dct_second_columns[name_column]], update_df[column])

            # Удаляем колонки с _y
        update_df.drop(columns=[column for column in update_df.columns if column.endswith('_y')], inplace=True)

        # Переименовываем колонки с _x
        update_df.columns = list(map(lambda x: x[:-2] if x.endswith('_x') else x, update_df.columns))

        # удаляем строки с _merge == right_only
        update_df = update_df[update_df['_merge'] != 'right_only']

        # Удаляем служебные колонки
        update_df.drop(columns=['ID_объединения', '_merge'], inplace=True)

        # используем уже созданный датафрейм right_df Удаляем лишнюю колонку в right_df
        right_df.drop(columns=['ID_объединения'], inplace=True)

        # Добавляем нехватающие колонки
        new_right_df = right_df.reindex(columns=update_df.columns, fill_value=None)

        update_df = pd.concat([update_df, new_right_df])

        for r in dataframe_to_rows(update_df, index=False, header=True):
            wb['Обновленная таблица'].append(r)

        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Сохраняем итоговый файл
        wb.save(f'{path_to_end_folder_comparison}/Результат слияния 2 таблиц от {current_time}.xlsx')
        # Сохраняем отдельно обновленную таблицу
        update_df.to_excel(
            f'{path_to_end_folder_comparison}/Таблица с обновленными данными и колонками от {current_time}.xlsx',
            index=False)

    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')
    except KeyError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В таблице нет такой колонки!\nПроверьте написание названия колонки')
        logging.exception('AN ERROR HAS OCCURRED')
    except ValueError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В таблице нет листа с таким названием!\nПроверьте написание названия листа')
        logging.exception('AN ERROR HAS OCCURRED')

    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Данные успешно обработаны')
