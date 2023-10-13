"""
Скрипт для разделения списка по значениям выбранной колонки.Результаты сохраняются либо в листы одного файла либо
в отдельные файлы. Например разделить большой список по полу или по группам
"""
import pandas as pd
import os
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from collections import Counter
import time
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)

def split_table(file_data_split:str,name_sheet:str,number_column:int,checkbox_split:int,path_to_end_folder):
    """
    Функция для разделения таблицы по значениям в определенном листе и колонке. Разделение по файлам и листам с сохранением названий

    :param file_data_split: файл с таблицей
    :param name_sheet: имя листа с таблицей
    :param number_column:порядковый номер колонки , прибавляется 1 чтобы соответстовать экселю
    :param checkbox_split: вариант разделения
    :param path_to_end_folder: путь к итоговой папке
    :return: один файл в котором много листов либо много файлов в зависимости от режима
    """
    df = pd.read_excel(file_data_split,sheet_name=name_sheet,dtype=str)
    lst_value_column = df.iloc[:,number_column-1].unique() # получаем все значения нужной колонки, -1 отнимаем поскольку в экселе нумерация с 1
    used_name_sheet = set() # множество для хранения значений которые уже были использованы

    name_column = df.columns[number_column-1] # получаем название колонки
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S',t)

    if checkbox_split == 0:
        wb = openpyxl.Workbook() # создаем файл
        for idx,value in enumerate(lst_value_column):
            temp_df = df[df[name_column] == value] # отфильтровываем по значению
            short_value = value[:20] # получаем обрезанное значение
            if short_value in used_name_sheet:
                short_value = f'{short_value}_{idx}' # добавляем окончание
            wb.create_sheet(short_value,index=idx) # создаем лист
            used_name_sheet.add(short_value)
            for row in dataframe_to_rows(temp_df,index=False,header=True):
                wb[short_value].append(row)

            # Устанавливаем автоширину для каждой колонки
            for column in wb[short_value].columns:
                max_length = 0
                column_name = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                wb[short_value].column_dimensions[column_name].width = adjusted_width





        wb.save(f'{path_to_end_folder}\Вариант А один файл {current_time}.xlsx')





if __name__ == '__main__':
    file_data = 'data/Разделение таблицы/Базовая таблица 1000 человек.xlsx'
    name_sheet_main = 'Sheet1'
    number_column_main = 16
    checkbox_split_main = 0
    path_to_end_folder_main = 'data/Разделение таблицы/result'

    split_table(file_data,name_sheet_main, number_column_main, checkbox_split_main, path_to_end_folder_main)
    print('Lindy Booth')



