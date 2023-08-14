"""
Графический интерфейс
"""
from diff_tables import find_diffrence # Функции для нахождения разницы 2 таблиц
from decl_case import declension_fio_by_case  # Функция для склонения ФИО по падежам
from comparsion_two_tables import merging_two_tables # Функция для сравнения, слияния 2 таблиц
from table_stat import counting_by_category # Функция для подсчета категориальных переменныъ
from table_stat import counting_quantitative_stat # функция для подсчета количественных статистик
from processing_date import proccessing_date # Функция для объединения двух таблиц
from union_tables import union_tables # Функция для объедения множества таблиц

import pandas as pd
import numpy as np
import os
from dateutil.parser import ParserError
from docxtpl import DocxTemplate
from docxcompose.composer import Composer
from docx import Document
from docx2pdf import convert
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import pytrovich
from pytrovich.detector import PetrovichGenderDetector
from pytrovich.enums import NamePart, Gender, Case
from pytrovich.maker import PetrovichDeclinationMaker
import time
import datetime
import warnings
from collections import Counter

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import sys
import locale
import logging
import tempfile
import re

logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)


# Классы для исключений

class CheckBoxException(Exception):
    """
    Класс для вызовы исключения в случае если неправильно выставлены чекбоксы
    """
    pass


class NotFoundValue(Exception):
    """
    Класс для обозначения того что значение не найдено
    """
    pass


class ShapeDiffierence(Exception):
    """
    Класс для обозначения несовпадения размеров таблицы
    """
    pass


class ColumnsDifference(Exception):
    """
    Класс для обозначения того что названия колонок не совпадают
    """
    pass


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_file_template_doc():
    """
    Функция для выбора файла шаблона
    :return: Путь к файлу шаблона
    """
    global name_file_template_doc
    name_file_template_doc = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))


def select_file_data_doc():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global name_file_data_doc
    # Получаем путь к файлу
    name_file_data_doc = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_doc():
    """
    Функция для выбора папки куда будут генерироваться файлы
    :return:
    """
    global path_to_end_folder_doc
    path_to_end_folder_doc = filedialog.askdirectory()


# Функции для вкладки извлечение данных
def select_file_params_calculate_data():
    """
    Функция для выбора файла c ячейками которые нужно подсчитать
    :return: Путь к файлу
    """
    global name_file_params_calculate_data
    name_file_params_calculate_data = filedialog.askopenfilename(
        filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_files_data_calculate_data():
    """
    Функция для выбора файлов с данными параметры из которых нужно подсчитать
    :return: Путь к файлам с данными
    """
    global names_files_calculate_data
    # Получаем путь к файлу
    names_files_calculate_data = filedialog.askopenfilenames(
        filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_calculate_data():
    """
    Функция для выбора папки куда будут генерироваться файл  с результатом подсчета и файл с проверочной инфомрацией
    :return:
    """
    global path_to_end_folder_calculate_data
    path_to_end_folder_calculate_data = filedialog.askdirectory()


def calculate_data():
    """
    Функция для подсчета данных из файлов
    :return:
    """
    try:
        count = 0
        count_errors = 0
        quantity_files = len(names_files_calculate_data)
        current_time = time.strftime('%H_%M_%S')
        # Состояние чекбокса
        mode_text = mode_text_value.get()

        # Получаем название обрабатываемого листа
        name_list_df = pd.read_excel(name_file_params_calculate_data, nrows=2)
        name_list = name_list_df['Значение'].loc[0]

        # Получаем количество листов в файле, на случай если название листа не совпадает с правильным
        quantity_list_in_file = name_list_df['Значение'].loc[1]

        # Получаем шаблон с данными, первую строку пропускаем, поскольку название обрабатываемого листа мы уже получили
        df = pd.read_excel(name_file_params_calculate_data, skiprows=2)

        # Создаем словарь параметров
        param_dict = dict()

        for row in df.itertuples():
            param_dict[row[1]] = row[2]
        # Создаем словарь для подсчета данных, копируя ключи из словаря параметров, значения в зависимости от способа обработки

        if mode_text == 'Yes':
            result_dct = {key: '' for key, value in param_dict.items()}
        else:
            result_dct = {key: 0 for key, value in param_dict.items()}

            # Создаем датафрейм для контроля процесса подсчета и заполняем словарь на основе которого будем делать итоговую таблицу

        check_df = pd.DataFrame(columns=param_dict.keys())
        # Вставляем колонку для названия файла
        check_df.insert(0, 'Название файла', '')
        for file in names_files_calculate_data:
            # Проверяем чтобы файл не был резервной копией.
            if '~$' in file:
                continue
            # Создаем словарь для создания строки которую мы будем добавлять в проверочный датафрейм
            new_row = dict()
            # Получаем  отбрасываем расширение файла
            full_name_file = file.split('.')[0]
            # Получаем имя файла  без пути
            name_file = full_name_file.split('/')[-1]
            try:

                new_row['Название файла'] = name_file

                wb = openpyxl.load_workbook(file)
                # Проверяем наличие листа
                if name_list in wb.sheetnames:
                    sheet = wb[name_list]
                # проверяем количество листов в файле.Если значение равно 1 то просто берем первый лист, иначе вызываем ошибку
                elif quantity_list_in_file == 1:
                    temp_name = wb.sheetnames[0]
                    sheet = wb[temp_name]
                else:
                    raise Exception
                for key, cell in param_dict.items():
                    result_dct[key] += check_data(sheet[cell].value, mode_text)
                    new_row[key] = sheet[cell].value

                temp_df = pd.DataFrame(new_row, index=['temp_index'])
                check_df = pd.concat([check_df, temp_df], ignore_index=True)
                # check_df = check_df.append(new_row, ignore_index=True)

                count += 1
            # Ловим исключения
            except Exception as err:
                count_errors += 1
                with open(f'{path_to_end_folder_calculate_data}/Необработанные файлы {current_time}.txt', 'a',
                          encoding='utf-8') as f:
                    f.write(f'Файл {name_file} не обработан!!!\n')

        check_df.to_excel(f'{path_to_end_folder_calculate_data}/Проверка вычисления {current_time}.xlsx', index=False)

        # Создание итоговой таблицы результатов подсчета

        finish_result = pd.DataFrame()

        finish_result['Наименование показателя'] = result_dct.keys()
        finish_result['Значение показателя'] = result_dct.values()
        # Проводим обработку в зависимости от значения переключателя

        # Получаем текущее время для того чтобы использовать в названии

        if mode_text == 'Yes':
            # Обрабатываем датафрейм считая текстовые данные
            count_text_df = count_text_value(finish_result)
            count_text_df.to_excel(
                f'{path_to_end_folder_calculate_data}/Подсчет текстовых значений {current_time}.xlsx')
        else:
            finish_result.to_excel(f'{path_to_end_folder_calculate_data}/Итоговые значения {current_time}.xlsx',
                                   index=False)

        if count_errors != 0:
            messagebox.showinfo('Веста Обработка таблиц и создание документов ver 1.35',
                                f'Обработка файлов завершена!\nОбработано файлов:  {count} из {quantity_files}\n Необработанные файлы указаны в файле {path_to_end_folder_calculate_data}/ERRORS {current_time}.txt ')
        else:
            messagebox.showinfo('Веста Обработка таблиц и создание документов ver 1.35',
                                f'Обработка файлов успешно завершена!\nОбработано файлов:  {count} из {quantity_files}')
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log')


# Функции для слияния таблиц

def select_end_folder_merger():
    """
    Функция для выбора папки куда будет генерироваться итоговый файл
    :return:
    """
    global path_to_end_folder_merger
    path_to_end_folder_merger = filedialog.askdirectory()


def select_folder_data_merger():
    """
    Функция для выбора папки где хранятся нужные файлы
    :return:
    """
    global dir_name
    dir_name = filedialog.askdirectory()


def select_params_file_merger():
    """
    Функция для выбора файла c ячейками которые нужно подсчитать
    :return: Путь к файлу
    """
    if group_rb_type_harvest.get() == 2:
        global params_harvest
        params_harvest = filedialog.askopenfilename(
            filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))
    else:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             'Выберите вариант слияния В и попробуйте снова ')


def select_standard_file_merger():
    """
    Функция для выбора файла c ячейками которые нужно подсчитать
    :return: Путь к файлу
    """
    global file_standard_merger
    file_standard_merger = filedialog.askopenfilename(
        filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def merge_tables():
    """
    Функция для слияния таблиц с одинаковой структурой в одну большую таблицу
    """
    # Звпускаем функцию обработки
    try:
        checkbox_harvest = group_rb_type_harvest.get() # получаем значение чекбокса
        merger_entry_skip_rows = merger_number_skip_rows.get() # получаем сколько строк занимает заголовок
        # проверяем значение чекбокса на случай
        if checkbox_harvest != 2:
            file_params = None
        else:
            file_params = params_harvest

        union_tables(checkbox_harvest,merger_entry_skip_rows,file_standard_merger,dir_name,path_to_end_folder_merger, file_params)
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                         f'Выберите папку с файлами,эталонный файл и папку куда будут генерироваться файлы')


def count_text_value(df):
    """
    Функция для подсчета количества вариантов того или иного показателя
    :param df: датафрейм с сырыми данными. Название показателя значение показателя(строка разделенная ;)
    :return: обработанный датафрейм с мультиндексом, где (Название показателя это индекс верхнего уровня, вариант показателя это индекс второго уровня а значение это сколько раз встречался
    этот вариант в обрабатываемых файлах)
    """
    data = dict()

    #
    for row in df.itertuples():
        value = row[2]
        if type(value) == float or type(value) == int:
            continue
        # Создаем список, разделяя строку по ;
        lst_value = row[2].split(';')[:-1]
        #     # Отрезаем последний элемент, поскольку это пустое значение
        temp_df = pd.DataFrame({'Value': lst_value})
        counts_series = temp_df['Value'].value_counts()
        # Делаем индекс колонкой и превращаем в обычную таблицу
        index_count_values = counts_series.reset_index()
        # Итерируемся по таблице.Это делается чтобы заполнить словарь на основе которого будет создаваться итоговая таблица
        for count_row in index_count_values.itertuples():
            # print(count_row)
            # Заполняем словарь
            data[(row[1], count_row[1])] = count_row[2]
    # Создаем на основе получившегося словаря таблицу
    out_df = pd.Series(data).to_frame().reset_index()
    out_df = out_df.set_index(['level_0', 'level_1'])
    out_df.index.names = ['Название показателя', 'Вариант показателя']
    out_df.rename(columns={0: 'Количество'}, inplace=True)
    return out_df


def check_data(cell, text_mode):
    """
    Функция для проверки значения ячейки. Для обработки пустых значений, строковых значений, дат
    :param cell: значение ячейки
    :return: 0 если значение ячейки не число
            число если значение ячейки число(ха звучит глуповато)
    думаю функция должна работать с дополнительным параметром, от которого будет зависеть подсчет значений навроде галочек или плюсов в анкетах или опросах.
    """
    # Проверяем режим работы. если текстовый, то просто складываем строки
    if text_mode == 'Yes':
        if cell is None:
            return ''
        else:
            temp_str = str(cell)
            return f'{temp_str};'
    # Если режим работы стандартный. Убрал подсчет строк и символов в числовом режиме, чтобы не запутывать.
    else:
        if cell is None:
            return 0
        if type(cell) == int:
            return cell
        elif type(cell) == float:
            return cell
        else:
            return 0


def generate_docs_other():
    """
    Функция для создания документов из произвольных таблиц(т.е. отличающихся от структуры базы данных Веста Обработка таблиц и создание документов ver 1.35)
    :return:
    """
    try:
        name_column = entry_name_column_data.get()
        name_type_file = entry_type_file.get()
        name_value_column = entry_value_column.get()

        # получаем состояние чекбокса создания pdf
        mode_pdf = mode_pdf_value.get()

        # Считываем данные
        # Добавил параметр dtype =str чтобы данные не преобразовались а использовались так как в таблице
        df = pd.read_excel(name_file_data_doc, dtype=str)
        # Заполняем Nan
        df.fillna(' ', inplace=True)
        lst_date_columns = []

        for idx, column in enumerate(df.columns):
            if 'дата' in column.lower():
                lst_date_columns.append(idx)

        # Конвертируем в пригодный строковый формат
        for i in lst_date_columns:
            df.iloc[:, i] = pd.to_datetime(df.iloc[:, i], errors='coerce', dayfirst=True)
            df.iloc[:, i] = df.iloc[:, i].apply(create_doc_convert_date)

        # Конвертируем датафрейм в список словарей
        data = df.to_dict('records')
        # Получаем состояние  чекбокса объединения файлов в один

        mode_combine = mode_combine_value.get()
        # Получаем состояние чекбокса создания индвидуального файла
        mode_group = mode_group_doc.get()

        # В зависимости от состояния чекбоксов обрабатываем файлы
        if mode_combine == 'No':
            if mode_group == 'No':
                # Создаем в цикле документы
                for idx, row in enumerate(data):
                    doc = DocxTemplate(name_file_template_doc)
                    context = row
                    # print(context)
                    doc.render(context)
                    # Сохраняенм файл
                    # получаем название файла и убираем недопустимые символы < > : " /\ | ? *
                    name_file = f'{name_type_file} {row[name_column]}'
                    name_file = re.sub(r'[<> :"?*|\\/]', ' ', name_file)
                    # проверяем файл на наличие, если файл с таким названием уже существует то добавляем окончание
                    if os.path.exists(f'{path_to_end_folder_doc}/{name_file}.docx'):
                        doc.save(f'{path_to_end_folder_doc}/{name_file}_{idx}.docx')

                    doc.save(f'{path_to_end_folder_doc}/{name_file}.docx')
                    if mode_pdf == 'Yes':
                        convert(f'{path_to_end_folder_doc}/{name_file}.docx',
                                f'{path_to_end_folder_doc}/{name_file}.pdf', keep_active=True)
            else:
                # Отбираем по значению строку

                single_df = df[df[name_column] == name_value_column]
                # Конвертируем датафрейм в список словарей
                single_data = single_df.to_dict('records')
                # Проверяем количество найденных совпадений
                # очищаем от запрещенных символов
                name_file = f'{name_type_file} {name_value_column}'
                name_file = re.sub(r'[<> :"?*|\\/]', ' ', name_file)
                if len(single_data) == 1:
                    for row in single_data:
                        doc = DocxTemplate(name_file_template_doc)
                        doc.render(row)
                        # Сохраняенм файл
                        doc.save(f'{path_to_end_folder_doc}/{name_file}.docx')
                        if mode_pdf == 'Yes':
                            convert(f'{path_to_end_folder_doc}/{name_file}.docx',
                                    f'{path_to_end_folder_doc}/{name_file}.pdf', keep_active=True)
                elif len(single_data) > 1:
                    for idx, row in enumerate(single_data):
                        doc = DocxTemplate(name_file_template_doc)
                        doc.render(row)
                        # Сохраняем файл

                        doc.save(f'{path_to_end_folder_doc}/{name_file}_{idx}.docx')
                        if mode_pdf == 'Yes':
                            convert(f'{path_to_end_folder_doc}/{name_file}_{idx}.docx',
                                    f'{path_to_end_folder_doc}/{name_file}_{idx}.pdf', keep_active=True)
                else:
                    raise NotFoundValue



        else:
            if mode_group == 'No':
                # Список с созданными файлами
                files_lst = []
                # Создаем временную папку
                with tempfile.TemporaryDirectory() as tmpdirname:
                    print('created temporary directory', tmpdirname)
                    # Создаем и сохраняем во временную папку созданные документы Word
                    for row in data:
                        doc = DocxTemplate(name_file_template_doc)
                        context = row
                        doc.render(context)
                        # Сохраняем файл
                        # очищаем от запрещенных символов
                        name_file = f'{row[name_column]}'
                        name_file = re.sub(r'[<> :"?*|\\/]', ' ', name_file)

                        doc.save(f'{tmpdirname}/{name_file}.docx')
                        # Добавляем путь к файлу в список
                        files_lst.append(f'{tmpdirname}/{name_file}.docx')
                    # Получаем базовый файл
                    main_doc = files_lst.pop(0)
                    # Запускаем функцию
                    combine_all_docx(main_doc, files_lst)
            else:
                raise CheckBoxException

    except NameError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
        logging.exception('AN ERROR HAS OCCURRED')
    except KeyError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             f'В таблице не найдена указанная колонка {e.args}')
    except PermissionError:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             f'Закройте все файлы Word созданные Вестой')
        logging.exception('AN ERROR HAS OCCURRED')
    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except CheckBoxException:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             f'Уберите галочку из чекбокса Поставьте галочку, если вам нужно создать один документ\nдля конкретного значения (например для определенного ФИО)'
                             )
    except NotFoundValue:
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             f'Указанное значение не найдено в выбранной колонке\nПроверьте наличие такого значения в таблице'
                             )
    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log')

    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов ver 1.35', 'Создание документов завершено!')


def check_date_columns(i, value):
    """
    Функция для проверки типа колонки. Необходимо найти колонки с датой
    :param i:
    :param value:
    :return:
    """
    try:
        itog = pd.to_datetime(str(value), infer_datetime_format=True)
    except:
        pass
    else:
        return i


def set_rus_locale():
    """
    Функция чтобы можно было извлечь русские названия месяцев
    """
    locale.setlocale(
        locale.LC_ALL,
        'rus_rus' if sys.platform == 'win32' else 'ru_RU.UTF-8')


def combine_all_docx(filename_master, files_lst):
    """
    Функция для объединения файлов Word взято отсюда
    https://stackoverflow.com/questions/24872527/combine-word-document-using-python-docx
    :param filename_master: базовый файл
    :param files_list: список с созданными файлами
    :return: итоговый файл
    """
    # получаем значение переключателя pdf
    mode_pdf = mode_pdf_value.get()
    # Получаем текущее время
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    number_of_sections = len(files_lst)
    # Открываем и обрабатываем базовый файл
    master = Document(filename_master)
    composer = Composer(master)
    # Перебираем и добавляем файлы к базовому
    for i in range(0, number_of_sections):
        doc_temp = Document(files_lst[i])
        composer.append(doc_temp)
    # Сохраняем файл
    composer.save(f"{path_to_end_folder_doc}/Объединеный файл от {current_time}.docx")
    if mode_pdf == 'Yes':
        convert(f"{path_to_end_folder_doc}/Объединеный файл от {current_time}.docx",
                f"{path_to_end_folder_doc}/Объединеный файл от {current_time}.pdf", keep_active=True)


def convert_date(cell):
    """
    Функция для конвертации даты в формате 1957-05-10 в формат 10.05.1957(строковый)
    """

    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date

    except TypeError:
        print(cell)
        messagebox.showerror('Веста Обработка таблиц и создание документов ver 1.35',
                             'Проверьте правильность заполнения ячеек с датой!!!')
        logging.exception('AN ERROR HAS OCCURRED')
        quit()


def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'
    except TypeError:
        return 'Не удалось конвертировать дату.Проверьте значение ячейки!!!'


def processing_date_column(df, lst_columns):
    """
    Функция для обработки столбцов с датами. конвертация в строку формата ДД.ММ.ГГГГ
    """
    # получаем первую строку
    first_row = df.iloc[0, lst_columns]

    lst_first_row = list(first_row)  # Превращаем строку в список
    lst_date_columns = []  # Создаем список куда будем сохранять колонки в которых находятся даты
    tupl_row = list(zip(lst_columns,
                        lst_first_row))  # Создаем список кортежей формата (номер колонки,значение строки в этой колонке)

    for idx, value in tupl_row:  # Перебираем кортеж
        result = check_date_columns(idx, value)  # проверяем является ли значение датой
        if result:  # если да то добавляем список порядковый номер колонки
            lst_date_columns.append(result)
        else:  # иначе проверяем следующее значение
            continue
    for i in lst_date_columns:  # Перебираем список с колонками дат, превращаем их в даты и конвертируем в нужный строковый формат
        df.iloc[:, i] = pd.to_datetime(df.iloc[:, i], errors='coerce', dayfirst=True)
        df.iloc[:, i] = df.iloc[:, i].apply(create_doc_convert_date)


"""
Функции для получения параметров обработки даты рождения
"""
def select_file_data_date():
    """
    Функция для выбора файла с данными для которого нужно разбить по категориям
    :return: Путь к файлу с данными
    """
    global name_file_data_date
    # Получаем путь к файлу
    name_file_data_date = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_end_folder_date():
    """
    Функция для выбора папки куда будет генерироваться итоговый файл
    :return:
    """
    global path_to_end_folder_date
    path_to_end_folder_date = filedialog.askdirectory()

def calculate_date():
    """
    Функция для разбиения по категориям, подсчета текущего возраста и выделения месяца,года
    :return:
    """
    try:
        raw_selected_date = entry_date.get()
        name_column = entry_name_column.get()
        # Устанавливаем русскую локаль
        set_rus_locale()
        proccessing_date(raw_selected_date,name_column,name_file_data_date,path_to_end_folder_date)
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файл с данными и папку куда будет генерироваться файл')


"""
Функции для подсчета статистик по таблице
"""
def select_file_data_groupby():
    """
    Функция для выбора файла с данными
    :return:
    """
    global name_file_data_groupby
    # Получаем путь к файлу
    name_file_data_groupby = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_groupby():
    """
    Функция для выбора папки куда будет генерироваться итоговый файл
    :return:
    """
    global path_to_end_folder_groupby
    path_to_end_folder_groupby = filedialog.askdirectory()

def groupby_category():
    """
    Подсчет категорий по всем колонкам таблицы
    """
    try:
        counting_by_category(name_file_data_groupby,path_to_end_folder_groupby)
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файл с данными и папку куда будет генерироваться файл')

def groupby_stat():
    """
    Подсчет категорий по всем колонкам таблицы
    """
    try:
        counting_quantitative_stat(name_file_data_groupby,path_to_end_folder_groupby)
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файл с данными и папку куда будет генерироваться файл')

# Функциия для слияния 2 таблиц
def select_file_params_comparsion():
    """
    Функция для выбора файла с параметрами колонок т.е. кокие колонки нужно обрабатывать
    :return:
    """
    global file_params
    file_params = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_first_comparison():
    """
    Функция для выбора  первого файла с данными которые нужно сравнить
    :return: Путь к файлу с данными
    """
    global name_first_file_comparison
    # Получаем путь к файлу
    name_first_file_comparison = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_second_comparison():
    """
    Функция для выбора  второго файла с данными которые нужно сравнить
    :return: Путь к файлу с данными
    """
    global name_second_file_comparison
    # Получаем путь к файлу
    name_second_file_comparison = filedialog.askopenfilename(
        filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_comparison():
    """
    Функция для выбора папки куда будет генерироваться итоговый файл
    :return:
    """
    global path_to_end_folder_comparison
    path_to_end_folder_comparison = filedialog.askdirectory()


def processing_comparison():
    """
    Функция для сравнения,слияния 2 таблиц
    :return:
    """
    # получаем названия листов
    try:
        first_sheet = entry_first_sheet_name.get()
        second_sheet = entry_second_sheet_name.get()

        merging_two_tables(file_params,first_sheet,second_sheet,name_first_file_comparison,name_second_file_comparison, path_to_end_folder_comparison)
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')

"""
Функции для склонения ФИО по падежам
"""
def select_data_decl_case():
    """
    Функция для файла с данными
    :return: Путь к файлу с данными
    """
    global data_decl_case
    # Получаем путь к файлу
    data_decl_case = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_decl_case():
    """
    Функия для выбора папки.Определенно вот это когда нибудь я перепишу на ООП
    :return:
    """
    global path_to_end_folder_decl_case
    path_to_end_folder_decl_case = filedialog.askdirectory()


def process_decl_case():
    """
    Функция для проведения склонения ФИО по падежам
    :return:
    """
    try:
        fio_column = decl_case_entry_fio.get()
        declension_fio_by_case(fio_column,data_decl_case,path_to_end_folder_decl_case)
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')

"""
Нахождения разницы 2 таблиц
Функции  получения параметров для find_diffrenece 
"""

def select_first_diffrence():
    """
    Функция для файла с данными
    :return: Путь к файлу с данными
    """
    global data_first_diffrence
    # Получаем путь к файлу
    data_first_diffrence = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_second_diffrence():
    """
    Функция для файла с данными
    :return: Путь к файлу с данными
    """
    global data_second_diffrence
    # Получаем путь к файлу
    data_second_diffrence = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def select_end_folder_diffrence():
    """
    Функия для выбора папки.Определенно вот это когда нибудь я перепишу на ООП
    :return:
    """
    global path_to_end_folder_diffrence
    path_to_end_folder_diffrence = filedialog.askdirectory()


def processing_diffrence():
    """
    Функция для получения названий листов и путей к файлам которые нужно сравнить
    :return:
    """
    # названия листов в таблицах
    try:
        first_sheet = entry_first_sheet_name_diffrence.get()
        second_sheet = entry_second_sheet_name_diffrence.get()
        # находим разницу
        find_diffrence(first_sheet, second_sheet, data_first_diffrence, data_second_diffrence, path_to_end_folder_diffrence)
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')


"""
Функции для создания контекстного меню(Копировать,вставить,вырезать)
"""


def make_textmenu(root):
    """
    Функции для контекстного меню( вырезать,копировать,вставить)
    взято отсюда https://gist.github.com/angeloped/91fb1bb00f1d9e0cd7a55307a801995f
    """
    # эта штука делает меню
    global the_menu
    the_menu = Menu(root, tearoff=0)
    the_menu.add_command(label="Вырезать")
    the_menu.add_command(label="Копировать")
    the_menu.add_command(label="Вставить")
    the_menu.add_separator()
    the_menu.add_command(label="Выбрать все")


def callback_select_all(event):
    """
    Функции для контекстного меню( вырезать,копировать,вставить)
    взято отсюда https://gist.github.com/angeloped/91fb1bb00f1d9e0cd7a55307a801995f
    """
    # select text after 50ms
    window.after(50, lambda: event.widget.select_range(0, 'end'))


def show_textmenu(event):
    """
    Функции для контекстного меню( вырезать,копировать,вставить)
    взято отсюда https://gist.github.com/angeloped/91fb1bb00f1d9e0cd7a55307a801995f
    """
    e_widget = event.widget
    the_menu.entryconfigure("Вырезать", command=lambda: e_widget.event_generate("<<Cut>>"))
    the_menu.entryconfigure("Копировать", command=lambda: e_widget.event_generate("<<Copy>>"))
    the_menu.entryconfigure("Вставить", command=lambda: e_widget.event_generate("<<Paste>>"))
    the_menu.entryconfigure("Выбрать все", command=lambda: e_widget.select_range(0, 'end'))
    the_menu.tk.call("tk_popup", the_menu, event.x_root, event.y_root)


if __name__ == '__main__':
    window = Tk()
    window.title('Веста Обработка таблиц и создание документов ver 1.35')
    window.geometry('774x860+700+100')
    window.resizable(False, False)
    # Добавляем контекстное меню в поля ввода
    make_textmenu(window)

    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку создания документов по шаблону
    tab_create_doc = ttk.Frame(tab_control)
    tab_control.add(tab_create_doc, text='Создание\nдокументов')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Создание документов
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_create_doc,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nГенерация документов по шаблону'
                           '\nДля корректной работы программмы уберите из таблицы объединенные ячейки'
                           '\nДанные обрабатываются только с первого листа файла Excel!!!')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')
    img = PhotoImage(file=path_to_img)
    Label(tab_create_doc,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_doc = LabelFrame(tab_create_doc, text='Подготовка')
    frame_data_for_doc.grid(column=0, row=2, padx=10)

    # Создаем кнопку Выбрать шаблон
    btn_template_doc = Button(frame_data_for_doc, text='1) Выберите шаблон документа', font=('Arial Bold', 15),
                              command=select_file_template_doc
                              )
    btn_template_doc.grid(column=0, row=3, padx=10, pady=10)
    #
    # Создаем кнопку Выбрать файл с данными
    btn_data_doc = Button(frame_data_for_doc, text='2) Выберите файл с данными', font=('Arial Bold', 15),
                          command=select_file_data_doc
                          )
    btn_data_doc.grid(column=0, row=4, padx=10, pady=10)
    #
    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    # Определяем текстовую переменную
    entry_name_column_data = StringVar()
    # Описание поля
    label_name_column_data = Label(frame_data_for_doc,
                                   text='3) Введите название колонки в таблице\n по которой будут создаваться имена файлов')
    label_name_column_data.grid(column=0, row=5, padx=10, pady=5)
    # поле ввода
    data_column_entry = Entry(frame_data_for_doc, textvariable=entry_name_column_data, width=30)
    data_column_entry.grid(column=0, row=6, padx=5, pady=5, ipadx=30, ipady=4)

    # Поле для ввода названия генериуемых документов
    # Определяем текстовую переменную
    entry_type_file = StringVar()
    # Описание поля
    label_name_column_type_file = Label(frame_data_for_doc, text='4) Введите название создаваемых документов')
    label_name_column_type_file.grid(column=0, row=7, padx=10, pady=5)
    # поле ввода
    type_file_column_entry = Entry(frame_data_for_doc, textvariable=entry_type_file, width=30)
    type_file_column_entry.grid(column=0, row=8, padx=5, pady=5, ipadx=30, ipady=4)

    btn_choose_end_folder_doc = Button(frame_data_for_doc, text='5) Выберите конечную папку', font=('Arial Bold', 15),
                                       command=select_end_folder_doc
                                       )
    btn_choose_end_folder_doc.grid(column=0, row=9, padx=10, pady=10)

    # Создаем область для того чтобы поместить туда опции
    frame_data_for_options = LabelFrame(tab_create_doc, text='Дополнительные опции')
    frame_data_for_options.grid(column=0, row=10, padx=10)

    # Создаем переменную для хранения результа переключения чекбокса
    mode_combine_value = StringVar()

    # Устанавливаем значение по умолчанию для этой переменной. По умолчанию будет вестись подсчет числовых данных
    mode_combine_value.set('No')
    # Создаем чекбокс для выбора режима подсчета

    chbox_mode_calculate = Checkbutton(frame_data_for_options,
                                       text='Поставьте галочку, если вам нужно чтобы все файлы были объединены в один',
                                       variable=mode_combine_value,
                                       offvalue='No',
                                       onvalue='Yes')
    chbox_mode_calculate.grid(column=0, row=11, padx=1, pady=1)

    # Создаем чекбокс для режима создания pdf
    # Создаем переменную для хранения результа переключения чекбокса
    mode_pdf_value = StringVar()

    # Устанавливаем значение по умолчанию для этой переменной. По умолчанию будет вестись подсчет числовых данных
    mode_pdf_value.set('No')
    # Создаем чекбокс для выбора режима подсчета

    chbox_mode_pdf = Checkbutton(frame_data_for_options,
                                 text='Поставьте галочку, если вам нужно чтобы \n'
                                      'дополнительно создавались pdf версии документов',
                                 variable=mode_pdf_value,
                                 offvalue='No',
                                 onvalue='Yes')
    chbox_mode_pdf.grid(column=0, row=12, padx=1, pady=1)

    # создаем чекбокс для единичного документа

    # Создаем переменную для хранения результа переключения чекбокса
    mode_group_doc = StringVar()

    # Устанавливаем значение по умолчанию для этой переменной. По умолчанию будет вестись подсчет числовых данных
    mode_group_doc.set('No')
    # Создаем чекбокс для выбора режима подсчета
    chbox_mode_group = Checkbutton(frame_data_for_options,
                                   text='Поставьте галочку, если вам нужно создать один документ\nдля конкретного значения (например для определенного ФИО)',
                                   variable=mode_group_doc,
                                   offvalue='No',
                                   onvalue='Yes')
    chbox_mode_group.grid(column=0, row=13, padx=1, pady=1)
    # Создаем поле для ввода значения по которому будет создаваться единичный документ
    # Определяем текстовую переменную
    entry_value_column = StringVar()
    # Описание поля
    label_name_column_group = Label(frame_data_for_options,
                                    text='Введите значение из колонки\nуказанной на шаге 3 для которого нужно создать один документ,\nнапример конкретное ФИО')
    label_name_column_group.grid(column=0, row=14, padx=1, pady=1)
    # поле ввода
    type_file_group_entry = Entry(frame_data_for_options, textvariable=entry_value_column, width=30)
    type_file_group_entry.grid(column=0, row=15, padx=5, pady=5, ipadx=30, ipady=4)

    # Создаем кнопку для создания документов из таблиц с произвольной структурой
    btn_create_files_other = Button(tab_create_doc, text='6) Создать документ(ы)',
                                    font=('Arial Bold', 15),
                                    command=generate_docs_other
                                    )
    btn_create_files_other.grid(column=0, row=14, padx=1, pady=1)

    # Создаем вклдаку для обработки дат рождения

    tab_calculate_date = ttk.Frame(tab_control)
    tab_control.add(tab_calculate_date, text='Обработка\nдат рождения')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Обработка дат рождения
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_calculate_date,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nПодсчет по категориям,выделение месяца,года,подсчет текущего возраста'
                           '\nДля корректной работы программмы уберите из таблицы объединенные ячейки'
                           '\nДанные обрабатываются только с первого листа файла Excel!!!')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')
    img_date = PhotoImage(file=path_to_img)
    Label(tab_calculate_date,
          image=img_date
          ).grid(column=1, row=0, padx=10, pady=25)

    # Определяем текстовую переменную которая будет хранить дату
    entry_date = StringVar()
    # Описание поля
    label_name_date_field = Label(tab_calculate_date,
                                  text='Введите  дату в формате XX.XX.XXXX\n относительно, которой нужно подсчитать текущий возраст')
    label_name_date_field.grid(column=0, row=2, padx=10, pady=10)
    # поле ввода
    date_field = Entry(tab_calculate_date, textvariable=entry_date, width=30)
    date_field.grid(column=0, row=3, padx=5, pady=5, ipadx=30, ipady=15)

    # Создаем кнопку Выбрать файл с данными
    btn_data_date = Button(tab_calculate_date, text='1) Выберите файл с данными', font=('Arial Bold', 20),
                           command=select_file_data_date)
    btn_data_date.grid(column=0, row=4, padx=10, pady=10)

    btn_choose_end_folder_date = Button(tab_calculate_date, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                        command=select_end_folder_date
                                        )
    btn_choose_end_folder_date.grid(column=0, row=5, padx=10, pady=10)

    # Определяем текстовую переменную
    entry_name_column = StringVar()
    # Описание поля
    label_name_column = Label(tab_calculate_date,
                              text='3) Введите название колонки с датами рождения,\nкоторые нужно обработать ')
    label_name_column.grid(column=0, row=6, padx=10, pady=10)
    # поле ввода
    column_entry = Entry(tab_calculate_date, textvariable=entry_name_column, width=30)
    column_entry.grid(column=0, row=7, padx=7, pady=5, ipadx=30, ipady=15)

    btn_calculate_date = Button(tab_calculate_date, text='4) Обработать', font=('Arial Bold', 20),
                                command=calculate_date)
    btn_calculate_date.grid(column=0, row=8, padx=10, pady=10)

    # Создаем вкладку для подсчета данных по категориям
    tab_groupby_data = ttk.Frame(tab_control)
    tab_control.add(tab_groupby_data, text='Подсчет\nданных')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Подсчет данных  по категориям
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_groupby_data,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nПодсчет данных'
                           '\nДанные обрабатываются только с первого листа файла Excel!!!')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')
    img_groupby = PhotoImage(file=path_to_img)
    Label(tab_groupby_data,
          image=img_groupby
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_groupby = LabelFrame(tab_groupby_data, text='Подготовка')
    frame_data_for_groupby.grid(column=0, row=2, padx=10)

    # Создаем кнопку Выбрать файл с данными
    btn_data_groupby = Button(frame_data_for_groupby, text='1) Выберите файл с данными', font=('Arial Bold', 20),
                              command=select_file_data_groupby
                              )
    btn_data_groupby.grid(column=0, row=3, padx=10, pady=10)

    btn_choose_end_folder_groupby = Button(frame_data_for_groupby, text='2) Выберите конечную папку',
                                           font=('Arial Bold', 20),
                                           command=select_end_folder_groupby
                                           )
    btn_choose_end_folder_groupby.grid(column=0, row=4, padx=10, pady=10)

    # Создаем кнопки подсчета

    btn_groupby_category = Button(tab_groupby_data, text='Подсчитать количество по категориям\nдля всех колонок',
                                  font=('Arial Bold', 20),
                                  command=groupby_category)
    btn_groupby_category.grid(column=0, row=5, padx=10, pady=10)

    btn_groupby_stat = Button(tab_groupby_data, text='Подсчитать базовую статистику\nдля всех колонок',
                              font=('Arial Bold', 20),
                              command=groupby_stat)
    btn_groupby_stat.grid(column=0, row=6, padx=10, pady=10)

    # Создаем вкладку для сравнения 2 столбцов

    tab_comparison = ttk.Frame(tab_control)
    tab_control.add(tab_comparison, text='Слияние\n2 таблиц')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Создание документов
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_comparison,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                           '\nДля корректной работы программмы уберите из таблицы объединенные ячейки')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_com = resource_path('logo.png')
    img_comparison = PhotoImage(file=path_com)
    Label(tab_comparison,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_comparison = LabelFrame(tab_comparison, text='Подготовка')
    frame_data_for_comparison.grid(column=0, row=2, padx=10)

    # Создаем кнопку выбрать файл с параметрами
    btn_columns_params = Button(frame_data_for_comparison, text='1) Выберите файл с параметрами слияния',
                                font=('Arial Bold', 10),
                                command=select_file_params_comparsion)
    btn_columns_params.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку Выбрать  первый файл с данными
    btn_data_first_comparison = Button(frame_data_for_comparison, text='2) Выберите первый файл с данными',
                                       font=('Arial Bold', 10),
                                       command=select_first_comparison
                                       )
    btn_data_first_comparison.grid(column=0, row=4, padx=10, pady=10)

    # Определяем текстовую переменную
    entry_first_sheet_name = StringVar()
    # Описание поля
    label_first_sheet_name = Label(frame_data_for_comparison,
                                   text='3) Введите название листа в первом файле')
    label_first_sheet_name.grid(column=0, row=5, padx=10, pady=10)
    # поле ввода имени листа
    first_sheet_name_entry = Entry(frame_data_for_comparison, textvariable=entry_first_sheet_name, width=30)
    first_sheet_name_entry.grid(column=0, row=6, padx=5, pady=5, ipadx=15, ipady=10)

    # Создаем кнопку Выбрать  второй файл с данными
    btn_data_second_comparison = Button(frame_data_for_comparison, text='4) Выберите второй файл с данными',
                                        font=('Arial Bold', 10),
                                        command=select_second_comparison
                                        )
    btn_data_second_comparison.grid(column=0, row=7, padx=10, pady=10)

    # Определяем текстовую переменную
    entry_second_sheet_name = StringVar()
    # Описание поля
    label_second_sheet_name = Label(frame_data_for_comparison,
                                    text='5) Введите название листа во втором файле')
    label_second_sheet_name.grid(column=0, row=8, padx=10, pady=10)
    # поле ввода
    second__sheet_name_entry = Entry(frame_data_for_comparison, textvariable=entry_second_sheet_name, width=30)
    second__sheet_name_entry.grid(column=0, row=9, padx=5, pady=5, ipadx=15, ipady=10)

    # Создаем кнопку выбора папки куда будет генерироваьться файл
    btn_select_end_comparison = Button(frame_data_for_comparison, text='6) Выберите конечную папку',
                                       font=('Arial Bold', 10),
                                       command=select_end_folder_comparison
                                       )
    btn_select_end_comparison.grid(column=0, row=10, padx=10, pady=10)

    # Создаем кнопку Обработать данные
    btn_data_do_comparison = Button(tab_comparison, text='7) Произвести слияние\nтаблиц', font=('Arial Bold', 20),
                                    command=processing_comparison
                                    )
    btn_data_do_comparison.grid(column=0, row=11, padx=10, pady=10)

    # Создаем вкладку для обработки таблиц excel  с одинаковой структурой
    tab_calculate_data = ttk.Frame(tab_control)
    tab_control.add(tab_calculate_data, text='Извлечение\nданных')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вклдаку Обработки данных
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_calculate_data,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nИзвлечение данных из файлов Excel\nс одинаковой структурой')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_calculate = resource_path('logo.png')
    img_calculate = PhotoImage(file=path_to_img)
    Label(tab_calculate_data,
          image=img_calculate
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с параметрами
    btn_select_file_params = Button(tab_calculate_data, text='1) Выбрать файл с параметрами', font=('Arial Bold', 20),
                                    command=select_file_params_calculate_data
                                    )
    btn_select_file_params.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку Выбрать файл с данными
    btn_select_files_data = Button(tab_calculate_data, text='2) Выбрать файлы с данными', font=('Arial Bold', 20),
                                   command=select_files_data_calculate_data
                                   )
    btn_select_files_data.grid(column=0, row=3, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder = Button(tab_calculate_data, text='3) Выбрать конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder_calculate_data
                                   )
    btn_choose_end_folder.grid(column=0, row=4, padx=10, pady=10)

    # Создаем переменную для хранения результа переключения чекбокса
    mode_text_value = StringVar()

    # Устанавливаем значение по умолчанию для этой переменной. По умолчанию будет вестись подсчет числовых данных
    mode_text_value.set('No')
    # Создаем чекбокс для выбора режима подсчета

    chbox_mode_calculate = Checkbutton(tab_calculate_data,
                                       text='Поставьте галочку, если вам нужно подсчитать текстовые данные ',
                                       variable=mode_text_value,
                                       offvalue='No',
                                       onvalue='Yes')
    chbox_mode_calculate.grid(column=0, row=5, padx=10, pady=10)

    # Создаем кнопку для запуска подсчета файлов

    btn_calculate = Button(tab_calculate_data, text='4) Подсчитать', font=('Arial Bold', 20),
                           command=calculate_data
                           )
    btn_calculate.grid(column=0, row=6, padx=10, pady=10)

    """
    Создание вкладки для объединения таблиц в одну большую
    """
    # Создаем вкладку для подсчета данных по категориям
    tab_merger_tables = ttk.Frame(tab_control)
    tab_control.add(tab_merger_tables, text='Слияние\nфайлов')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Подсчет данных  по категориям
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_merger_tables,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nСлияние файлов Excel с одинаковой структурой'
                           '\nДля корректной работы программмы уберите из таблицы объединенные ячейки'
                      )
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')
    img_merger = PhotoImage(file=path_to_img)
    Label(tab_merger_tables,
          image=img_merger
          ).grid(column=1, row=0, padx=10, pady=25)

    # Переключатель:вариант слияния файлов
    # Создаем переключатель
    group_rb_type_harvest = IntVar()
    # Создаем фрейм для размещения переключателей(pack и грид не используются в одном контейнере)
    frame_rb_type_harvest = LabelFrame(tab_merger_tables, text='1) Выберите вариант слияния')
    frame_rb_type_harvest.grid(column=0, row=1, padx=10)
    #
    Radiobutton(frame_rb_type_harvest, text='А) Простое слияние по названию листов', variable=group_rb_type_harvest,
                value=0).pack()
    Radiobutton(frame_rb_type_harvest, text='Б) Слияние по порядку листов', variable=group_rb_type_harvest,
                value=1).pack()
    Radiobutton(frame_rb_type_harvest, text='В) Сложное слияние по названию листов', variable=group_rb_type_harvest,
                value=2).pack()

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_merger = LabelFrame(tab_merger_tables, text='Подготовка')
    frame_data_for_merger.grid(column=0, row=2, padx=10)

    # Создаем кнопку Выбрать папку с данными

    btn_data_merger = Button(frame_data_for_merger, text='2) Выберите папку с данными', font=('Arial Bold', 14),
                             command=select_folder_data_merger
                             )
    btn_data_merger.grid(column=0, row=3, padx=5, pady=5)

    # Создаем кнопку Выбрать эталонный файл

    btn_example_merger = Button(frame_data_for_merger, text='3) Выберите эталонный файл', font=('Arial Bold', 14),
                                command=select_standard_file_merger)
    btn_example_merger.grid(column=0, row=4, padx=5, pady=5)

    btn_choose_end_folder_merger = Button(frame_data_for_merger, text='4) Выберите конечную папку',
                                          font=('Arial Bold', 14),
                                          command=select_end_folder_merger
                                          )
    btn_choose_end_folder_merger.grid(column=0, row=5, padx=5, pady=5)

    # Определяем переменную в которой будем хранить количество пропускаемых строк
    merger_entry_skip_rows = StringVar()
    # Описание поля
    merger_label_skip_rows = Label(frame_data_for_merger,
                                   text='5) Введите количество строк\nв листах,чтобы пропустить\nзаголовок\n'
                                        'ТОЛЬКО для вариантов слияния А и Б ')
    merger_label_skip_rows.grid(column=0, row=8, padx=10, pady=10)
    # поле ввода
    merger_number_skip_rows = Entry(frame_data_for_merger, textvariable=merger_entry_skip_rows, width=5)
    merger_number_skip_rows.grid(column=0, row=9, padx=5, pady=5, ipadx=10, ipady=7)

    # Создаем кнопку выбора файла с параметрами
    btn_params_merger = Button(frame_data_for_merger, text='Выберите файл с параметрами слияния\n'
                                                           'ТОЛЬКО для варианта В', font=('Arial Bold', 14),
                               command=select_params_file_merger)
    btn_params_merger.grid(column=0, row=10, padx=5, pady=5)
    # Создаем кнопку слияния

    btn_merger_process = Button(tab_merger_tables, text='6) Произвести слияние \nфайлов',
                                font=('Arial Bold', 20),
                                command=merge_tables)
    btn_merger_process.grid(column=0, row=11, padx=10, pady=10)

    """
    Создание вкладки для склонения ФИО по падежам
    """
    # Создаем вкладку для подсчета данных по категориям
    tab_decl_by_cases = ttk.Frame(tab_control)
    tab_control.add(tab_decl_by_cases, text='Склонение ФИО\nпо падежам')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку Подсчет данных  по категориям
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_decl_by_cases,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\nСклонение ФИО по падежам и создание инициалов'
                           '\nДля корректной работы программмы уберите из таблицы объединенные ячейки'
                      )
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img = resource_path('logo.png')
    img_decl_by_cases = PhotoImage(file=path_to_img)
    Label(tab_decl_by_cases,
          image=img_decl_by_cases
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_decl_case = LabelFrame(tab_decl_by_cases, text='Подготовка')
    frame_data_for_decl_case.grid(column=0, row=2, padx=10)

    # выбрать файл с данными
    btn_data_decl_case = Button(frame_data_for_decl_case, text='1) Выберите файл с данными', font=('Arial Bold', 20),
                                command=select_data_decl_case)
    btn_data_decl_case.grid(column=0, row=3, padx=10, pady=10)

    # Ввести название колонки с ФИО
    # # Определяем переменную
    decl_case_fio_col = StringVar()
    # Описание поля ввода
    decl_case_label_fio = Label(frame_data_for_decl_case,
                                text='2) Введите название колонки\n с ФИО в им.падеже')
    decl_case_label_fio.grid(column=0, row=4, padx=10, pady=10)
    # поле ввода
    decl_case_entry_fio = Entry(frame_data_for_decl_case, textvariable=decl_case_fio_col, width=25)
    decl_case_entry_fio.grid(column=0, row=5, padx=5, pady=5, ipadx=10, ipady=7)
    #
    btn_choose_end_folder_decl_case = Button(frame_data_for_decl_case, text='3) Выберите конечную папку',
                                             font=('Arial Bold', 20),
                                             command=select_end_folder_decl_case
                                             )
    btn_choose_end_folder_decl_case.grid(column=0, row=6, padx=10, pady=10)

    # Создаем кнопку склонения по падежам

    btn_decl_case_process = Button(tab_decl_by_cases, text='4) Произвести склонение \nпо падежам',
                                   font=('Arial Bold', 20),
                                   command=process_decl_case)
    btn_decl_case_process.grid(column=0, row=7, padx=10, pady=10)

    """
    Разница двух таблиц
    """
    tab_diffrence = ttk.Frame(tab_control)
    tab_control.add(tab_diffrence, text='Разница\n2 таблиц')
    tab_control.pack(expand=1, fill='both')

    # Добавляем виджеты на вкладку разница 2 двух таблиц
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_diffrence,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                           'Количество строк и колонок в таблицах должно совпадать\n'
                           'Названия колонок в таблицах должны совпадать'
                           '\nДля корректной работы программмы уберите из таблицы объединенные ячейки')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_com = resource_path('logo.png')
    img_diffrence = PhotoImage(file=path_com)
    Label(tab_diffrence,
          image=img
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем область для того чтобы поместить туда подготовительные кнопки(выбрать файл,выбрать папку и т.п.)
    frame_data_for_diffrence = LabelFrame(tab_diffrence, text='Подготовка')
    frame_data_for_diffrence.grid(column=0, row=2, padx=10)

    # Создаем кнопку Выбрать  первый файл с данными
    btn_data_first_diffrence = Button(frame_data_for_diffrence, text='1) Выберите файл с первой таблицей',
                                      font=('Arial Bold', 10),
                                      command=select_first_diffrence
                                      )
    btn_data_first_diffrence.grid(column=0, row=3, padx=10, pady=10)

    # Определяем текстовую переменную
    entry_first_sheet_name_diffrence = StringVar()
    # Описание поля
    label_first_sheet_name_diffrence = Label(frame_data_for_diffrence,
                                             text='2) Введите название листа, где находится первая таблица')
    label_first_sheet_name_diffrence.grid(column=0, row=4, padx=10, pady=10)
    # поле ввода имени листа
    first_sheet_name_entry_diffrence = Entry(frame_data_for_diffrence, textvariable=entry_first_sheet_name_diffrence,
                                             width=30)
    first_sheet_name_entry_diffrence.grid(column=0, row=5, padx=5, pady=5, ipadx=15, ipady=10)

    # Создаем кнопку Выбрать  второй файл с данными
    btn_data_second_diffrence = Button(frame_data_for_diffrence, text='3) Выберите файл со второй таблицей',
                                       font=('Arial Bold', 10),
                                       command=select_second_diffrence
                                       )
    btn_data_second_diffrence.grid(column=0, row=6, padx=10, pady=10)

    # Определяем текстовую переменную
    entry_second_sheet_name_diffrence = StringVar()
    # Описание поля
    label_second_sheet_name_diffrence = Label(frame_data_for_diffrence,
                                              text='4) Введите название листа, где находится вторая таблица')
    label_second_sheet_name_diffrence.grid(column=0, row=7, padx=10, pady=10)
    # поле ввода
    second__sheet_name_entry_diffrence = Entry(frame_data_for_diffrence, textvariable=entry_second_sheet_name_diffrence,
                                               width=30)
    second__sheet_name_entry_diffrence.grid(column=0, row=8, padx=5, pady=5, ipadx=15, ipady=10)

    # Создаем кнопку выбора папки куда будет генерироваьться файл
    btn_select_end_diffrence = Button(frame_data_for_diffrence, text='5) Выберите конечную папку',
                                      font=('Arial Bold', 10),
                                      command=select_end_folder_diffrence
                                      )
    btn_select_end_diffrence.grid(column=0, row=10, padx=10, pady=10)

    # Создаем кнопку Обработать данные
    btn_data_do_diffrence = Button(tab_diffrence, text='6) Обработать таблицы', font=('Arial Bold', 20),
                                   command=processing_diffrence
                                   )
    btn_data_do_diffrence.grid(column=0, row=11, padx=10, pady=10)
    window.bind_class("Entry", "<Button-3><ButtonRelease-3>", show_textmenu)
    window.bind_class("Entry", "<Control-a>", callback_select_all)
    window.mainloop()
